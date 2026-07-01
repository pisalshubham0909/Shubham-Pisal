import streamlit as st
import pandas as pd
import numpy as np
import io
import json
from parser import parse_gstr2b_json, parse_purchase_register, auto_detect_columns
from reconciliation import reconcile_data, generate_supplier_summary
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Page Config
st.set_page_config(
    page_title="GSTR-2B Reconciliation Tool",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom Premium Styling (Dark-Neutral Theme with Glassmorphism highlights)
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Plus Jakarta Sans', sans-serif;
    }
    
    h1, h2, h3, h4, h5, h6 {
        font-family: 'Outfit', sans-serif;
        font-weight: 700;
        letter-spacing: -0.02em;
    }
    
    /* Premium KPI Card Styling */
    .kpi-card {
        background: linear-gradient(135deg, rgba(30, 41, 59, 0.9) 0%, rgba(15, 23, 42, 0.95) 100%);
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 16px;
        padding: 24px;
        margin-bottom: 20px;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.15);
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    .kpi-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 30px rgba(0, 0, 0, 0.25);
        border-color: rgba(99, 102, 241, 0.2);
    }
    .kpi-title {
        color: #94A3B8;
        font-size: 0.85rem;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        font-weight: 600;
        margin-bottom: 8px;
    }
    .kpi-value {
        color: #F8FAFC;
        font-size: 1.85rem;
        font-weight: 700;
        font-family: 'Outfit', sans-serif;
    }
    .kpi-subtitle {
        color: #64748B;
        font-size: 0.75rem;
        margin-top: 6px;
    }
    
    /* Custom Badges */
    .badge {
        padding: 4px 10px;
        border-radius: 6px;
        font-size: 0.75rem;
        font-weight: 600;
        display: inline-block;
        text-transform: uppercase;
        letter-spacing: 0.03em;
    }
    .badge-matched { background-color: rgba(16, 185, 129, 0.15); color: #10B981; border: 1px solid rgba(16, 185, 129, 0.2); }
    .badge-fuzzy { background-color: rgba(6, 182, 212, 0.15); color: #06B6D4; border: 1px solid rgba(6, 182, 212, 0.2); }
    .badge-mismatch { background-color: rgba(245, 158, 11, 0.15); color: #F59E0B; border: 1px solid rgba(245, 158, 11, 0.2); }
    .badge-books { background-color: rgba(244, 63, 94, 0.15); color: #F43F5E; border: 1px solid rgba(244, 63, 94, 0.2); }
    .badge-gstr2b { background-color: rgba(99, 102, 241, 0.15); color: #6366F1; border: 1px solid rgba(99, 102, 241, 0.2); }
    
    /* Header decoration */
    .header-gradient {
        background: linear-gradient(90deg, #F8FAFC 0%, #C7D2FE 50%, #818CF8 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    
    /* Style tables */
    .dataframe {
        font-size: 0.85rem !important;
    }
    </style>
""", unsafe_allow_html=True)

# Generate Synthetic Data function including new GST edge cases (RCM, ISD, SEZ, Late files)
def get_synthetic_data():
    # 1. Create GSTR-2B Data
    gstr2b_json_content = {
        "gstin": "27MYCOMPANY123Z0",
        "rtnprd": "032026",
        "b2b": [
            {
                "ctin": "27ABCDE1234F1Z5",
                "lglNm": "Alpha Suppliers Ltd",
                "inv": [
                    {
                        "inum": "INV-101",
                        "idt": "10-03-2026",
                        "val": 118000.0,
                        "pos": "27",
                        "rchrg": "N",
                        "inv_typ": "R",
                        "itcelg": "Y",
                        "flddt": "11-04-2026",
                        "g3bfil": "Y",
                        "items": [{"num": 1, "rt": 18.0, "txval": 100000.0, "igst": 0.0, "cgst": 9000.0, "sgst": 9000.0, "cess": 0.0}]
                    }
                ]
            },
            {
                "ctin": "27FGHIJ5678K2Z9",
                "lglNm": "Beta Enterprise",
                "inv": [
                    {
                        "inum": "INV-201",
                        "idt": "15-03-2026",
                        "val": 23600.0,
                        "pos": "27",
                        "rchrg": "N",
                        "inv_typ": "R",
                        "itcelg": "Y",
                        "flddt": "10-04-2026",
                        "g3bfil": "Y",
                        "items": [{"num": 1, "rt": 18.0, "txval": 20000.0, "igst": 3600.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}]
                    },
                    {
                        "inum": "INV-202",
                        "idt": "22-03-2026",
                        "val": 11800.0,
                        "pos": "27",
                        "rchrg": "N",
                        "inv_typ": "R",
                        "itcelg": "Y",
                        "flddt": "15-04-2026",
                        "g3bfil": "N",  # GSTR-3B NOT FILED
                        "items": [{"num": 1, "rt": 18.0, "txval": 10000.0, "igst": 0.0, "cgst": 900.0, "sgst": 900.0, "cess": 0.0}]
                    }
                ]
            },
            {
                "ctin": "27KLMNO9012L3Z3",
                "lglNm": "Gamma Corp",
                "inv": [
                    {
                        "inum": "INV-301",
                        "idt": "25-03-2026",
                        "val": 118000.0,
                        "pos": "27",
                        "rchrg": "N",
                        "inv_typ": "R",
                        "itcelg": "N",  # Ineligible blocked ITC 17(5)
                        "flddt": "09-04-2026",
                        "g3bfil": "Y",
                        "items": [{"num": 1, "rt": 18.0, "txval": 100000.0, "igst": 0.0, "cgst": 9000.0, "sgst": 9000.0, "cess": 0.0}]
                    },
                    {
                        "inum": "INV-302",
                        "idt": "28-03-2026",
                        "val": 17700.0,
                        "pos": "27",
                        "rchrg": "N",
                        "inv_typ": "R",
                        "itcelg": "Y",  # Target for fuzzy matching (books has "INV302-A")
                        "flddt": "11-04-2026",
                        "g3bfil": "Y",
                        "items": [{"num": 1, "rt": 18.0, "txval": 15000.0, "igst": 2700.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}]
                    }
                ]
            },
            {
                "ctin": "27PQRST3456M4Z1",
                "lglNm": "Delta Logistics",
                "inv": [
                    {
                        "inum": "INV-401",
                        "idt": "29-03-2026",
                        "val": 11800.0,
                        "pos": "27",
                        "rchrg": "Y",  # Reverse charge transaction
                        "inv_typ": "R",
                        "itcelg": "Y",
                        "flddt": "10-04-2026",
                        "g3bfil": "Y",
                        "items": [{"num": 1, "rt": 18.0, "txval": 10000.0, "igst": 0.0, "cgst": 900.0, "sgst": 900.0, "cess": 0.0}]
                    }
                ]
            }
        ],
        "b2ba": [
            {
                "ctin": "27ABCDE1234F1Z5",
                "lglNm": "Alpha Suppliers Ltd",
                "inv": [
                    {
                        "inum": "INV-99A",
                        "idt": "15-03-2026",
                        "oinum": "INV-99",
                        "oidt": "10-02-2026",
                        "val": 23600.0,
                        "pos": "27",
                        "rchrg": "N",
                        "inv_typ": "R",
                        "itcelg": "Y",
                        "flddt": "11-04-2026",
                        "g3bfil": "Y",
                        "items": [{"num": 1, "rt": 18.0, "txval": 20000.0, "igst": 3600.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}]
                    }
                ]
            }
        ],
        "cdnr": [
            {
                "ctin": "27FGHIJ5678K2Z9",
                "lglNm": "Beta Enterprise",
                "nt": [
                    {
                        "nt_num": "CN-01",
                        "nt_dt": "20-03-2026",
                        "val": 5900.0,
                        "nt_ty": "C",
                        "inum": "INV-201",
                        "idt": "15-03-2026",
                        "itcelg": "Y",
                        "pos": "27",
                        "rchrg": "N",
                        "flddt": "12-04-2026",
                        "g3bfil": "Y",
                        "items": [{"num": 1, "rt": 18.0, "txval": 5000.0, "igst": 900.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}]
                    }
                ]
            }
        ],
        "isd": [
            {
                "ctin": "27ISDHO9999A1Z2",
                "lglNm": "HO Head Office (ISD)",
                "doclist": [
                    {
                        "docnum": "ISD-88",
                        "docdt": "13-03-2026",
                        "val": 18000.0,
                        "itcelg": "Y",
                        "pos": "27",
                        "flddt": "13-04-2026",
                        "g3bfil": "Y",
                        "items": [{"num": 1, "rt": 0.0, "txval": 18000.0, "igst": 18000.0, "cgst": 0.0, "sgst": 0.0, "cess": 0.0}]
                    }
                ]
            }
        ],
        "impgsez": [
            {
                "boe_num": "BOE-SEZ-99",
                "boe_dt": "19-03-2026",
                "boe_val": 177000.0,
                "txval": 150000.0,
                "igst": 27000.0,
                "cess": 0.0,
                "itcelg": "Y",
                "ctin": "27SEZDV8888B3Z4",
                "lglNm": "SEZ Infrastructure Ltd",
                "pos": "27"
            }
        ],
        "impg": [
            {
                "boe_num": "BOE-501",
                "boe_dt": "18-03-2026",
                "boe_val": 500000.0,
                "port_cd": "INBOM4",
                "txval": 400000.0,
                "igst": 72000.0,
                "cess": 0.0,
                "itcelg": "Y"
            }
        ]
    }

    # 2. Create Books Data
    books_data = [
        # B2B Matched
        {"Supplier GSTIN": "27ABCDE1234F1Z5", "Supplier Name": "Alpha Suppliers Ltd", "Invoice Number": "INV-101", "Invoice Date": "10-03-2026", "Voucher Type": "Purchase", "Taxable Value": 100000.0, "IGST": 0.0, "CGST": 9000.0, "SGST": 9000.0, "POS": "27", "RCM": "No"},
        # Amended Invoice Match
        {"Supplier GSTIN": "27ABCDE1234F1Z5", "Supplier Name": "Alpha Suppliers Ltd", "Invoice Number": "INV-99", "Invoice Date": "10-02-2026", "Voucher Type": "Purchase", "Taxable Value": 20000.0, "IGST": 3600.0, "CGST": 0.0, "SGST": 0.0, "POS": "27", "RCM": "No"},
        # Rounded value mismatch within tolerance (₹2 difference)
        {"Supplier GSTIN": "27FGHIJ5678K2Z9", "Supplier Name": "Beta Enterprise", "Invoice Number": "INV-201", "Invoice Date": "15-03-2026", "Voucher Type": "Purchase", "Taxable Value": 20002.0, "IGST": 3600.36, "CGST": 0.0, "SGST": 0.0, "POS": "27", "RCM": "No"},
        # GSTR-3B Unfiled check
        {"Supplier GSTIN": "27FGHIJ5678K2Z9", "Supplier Name": "Beta Enterprise", "Invoice Number": "INV-202", "Invoice Date": "22-03-2026", "Voucher Type": "Purchase", "Taxable Value": 10000.0, "IGST": 0.0, "CGST": 900.0, "SGST": 900.0, "POS": "27", "RCM": "No"},
        # Only in books invoice
        {"Supplier GSTIN": "27FGHIJ5678K2Z9", "Supplier Name": "Beta Enterprise", "Invoice Number": "INV-203", "Invoice Date": "24-03-2026", "Voucher Type": "Purchase", "Taxable Value": 30000.0, "IGST": 5400.0, "CGST": 0.0, "SGST": 0.0, "POS": "27", "RCM": "No"},
        # Blocked ITC Match
        {"Supplier GSTIN": "27KLMNO9012L3Z3", "Supplier Name": "Gamma Corp", "Invoice Number": "INV-301", "Invoice Date": "25-03-2026", "Voucher Type": "Purchase", "Taxable Value": 100000.0, "IGST": 0.0, "CGST": 9000.0, "SGST": 9000.0, "POS": "27", "RCM": "No"},
        # Fuzzy match target: books has "INV302-A", JSON has "INV-302"
        {"Supplier GSTIN": "27KLMNO9012L3Z3", "Supplier Name": "Gamma Corp", "Invoice Number": "INV302-A", "Invoice Date": "28-03-2026", "Voucher Type": "Purchase", "Taxable Value": 15000.0, "IGST": 2700.0, "CGST": 0.0, "SGST": 0.0, "POS": "27", "RCM": "No"},
        # RCM match
        {"Supplier GSTIN": "27PQRST3456M4Z1", "Supplier Name": "Delta Logistics", "Invoice Number": "INV-401", "Invoice Date": "29-03-2026", "Voucher Type": "Purchase", "Taxable Value": 10000.0, "IGST": 0.0, "CGST": 900.0, "SGST": 900.0, "POS": "27", "RCM": "Yes"},
        # Credit Note
        {"Supplier GSTIN": "27FGHIJ5678K2Z9", "Supplier Name": "Beta Enterprise", "Invoice Number": "CN-01", "Invoice Date": "20-03-2026", "Voucher Type": "Credit Note", "Taxable Value": 5000.0, "IGST": 900.0, "CGST": 0.0, "SGST": 0.0, "POS": "27", "RCM": "No"},
        # ISD distribution match
        {"Supplier GSTIN": "27ISDHO9999A1Z2", "Supplier Name": "HO Head Office (ISD)", "Invoice Number": "ISD-88", "Invoice Date": "13-03-2026", "Voucher Type": "ISD Journal", "Taxable Value": 18000.0, "IGST": 18000.0, "CGST": 0.0, "SGST": 0.0, "POS": "27", "RCM": "No"},
        # SEZ Import match
        {"Supplier GSTIN": "27SEZDV8888B3Z4", "Supplier Name": "SEZ Infrastructure Ltd", "Invoice Number": "BOE-SEZ-99", "Invoice Date": "19-03-2026", "Voucher Type": "Import SEZ", "Taxable Value": 150000.0, "IGST": 27000.0, "CGST": 0.0, "SGST": 0.0, "POS": "27", "RCM": "No"},
        # Import matching
        {"Supplier GSTIN": "IMPORT", "Supplier Name": "Import of Goods", "Invoice Number": "BOE-501", "Invoice Date": "18-03-2026", "Voucher Type": "Import BOE", "Taxable Value": 400000.0, "IGST": 72000.0, "CGST": 0.0, "SGST": 0.0, "POS": "97", "RCM": "No"}
    ]
    df_books = pd.DataFrame(books_data)
    
    return gstr2b_json_content, df_books

# openpyxl Styled Excel Exporter to a Single Consolidated Sheet + Dashboard + Supplier Summary
def export_reco_to_excel(df_reco, df_supplier, summary_stats):
    wb = Workbook()
    
    # Styles
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    # Status fills
    matched_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    mismatch_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
    books_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    gstr2b_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
    
    status_fills = {
        'Matched': matched_fill,
        'Fuzzy Match': matched_fill,
        'Matched (Amended)': matched_fill,
        'Value Mismatch': mismatch_fill,
        'Date Mismatch': mismatch_fill,
        'Date & Value Mismatch': mismatch_fill,
        'Value Mismatch (Amended)': mismatch_fill,
        'Only in Books': books_fill,
        'Only in GSTR-2B': gstr2b_fill
    }

    white_bold_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    total_border = Border(
        top=Side(style='thin', color='94A3B8'),
        bottom=Side(style='double', color='0F172A')
    )

    # 1. TAB: Dashboard Overview
    ws_dash = wb.active
    ws_dash.title = "Dashboard Summary"
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash['A1'] = "GSTR-2B Reconciliation Report Dashboard"
    ws_dash['A1'].font = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    
    ws_dash['A3'] = "Reconciliation Metrics Summary"
    ws_dash['A3'].font = Font(name="Calibri", size=12, bold=True)
    
    dash_headers = ["Metric", "Books Value (INR)", "GSTR-2B Value (INR)", "Variance (INR)"]
    for col_num, header in enumerate(dash_headers, 1):
        cell = ws_dash.cell(row=4, column=col_num, value=header)
        cell.font = white_bold_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center")
        
    metrics_list = [
        ("Total Invoices Processed", summary_stats['books_count'], summary_stats['gstr2b_count'], summary_stats['books_count'] - summary_stats['gstr2b_count']),
        ("Total Taxable Value", summary_stats['books_taxable_total'], summary_stats['gstr2b_taxable_total'], summary_stats['taxable_variance']),
        ("Total Integrated Tax (IGST)", summary_stats['books_igst_total'], summary_stats['gstr2b_igst_total'], summary_stats['igst_variance']),
        ("Total Central Tax (CGST)", summary_stats['books_cgst_total'], summary_stats['gstr2b_cgst_total'], summary_stats['cgst_variance']),
        ("Total State Tax (SGST)", summary_stats['books_sgst_total'], summary_stats['gstr2b_sgst_total'], summary_stats['sgst_variance']),
        ("Total Eligible ITC Claimable", summary_stats['claimable_itc_total'], summary_stats['gstr2b_itc_total'], summary_stats['claimable_itc_total'] - summary_stats['gstr2b_itc_total']),
    ]
    
    for row_idx, item in enumerate(metrics_list, 5):
        for col_idx, val in enumerate(item, 1):
            cell = ws_dash.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx > 1:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                
    ws_dash.column_dimensions['A'].width = 30
    ws_dash.column_dimensions['B'].width = 25
    ws_dash.column_dimensions['C'].width = 25
    ws_dash.column_dimensions['D'].width = 25

    # Helper function to write sheet data
    def write_sheet(wb, title, data_df, is_reconciliation_report=False):
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True
        
        # Write Headers
        for col_num, col_name in enumerate(data_df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name.replace('_', ' ').title())
            cell.font = white_bold_font
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        # Write rows
        for row_num, (_, row_data) in enumerate(data_df.iterrows(), 2):
            for col_num, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # Format Dates
                if isinstance(val, pd.Timestamp) or (isinstance(val, str) and '-' in val and len(val) == 10):
                    if isinstance(val, pd.Timestamp):
                        cell.value = val.strftime('%d-%m-%Y') if not pd.isna(val) else ""
                    else:
                        cell.value = val
                    cell.alignment = Alignment(horizontal="center")
                # Format Numbers
                elif isinstance(val, (int, float, np.integer, np.floating)):
                    cell.value = float(val) if not pd.isna(val) else 0.0
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = "" if pd.isna(val) else str(val)
                    cell.alignment = Alignment(horizontal="left")
                
                cell.font = regular_font
                cell.border = thin_border
                
                # Highlight status column in reconciliation report
                if is_reconciliation_report and col_num == 1: # reco_status
                    status_val = str(val)
                    if status_val in status_fills:
                        cell.fill = status_fills[status_val]
                        
        # Add total row for numeric columns in reconciliation sheets
        if is_reconciliation_report and not data_df.empty:
            tot_row = len(data_df) + 2
            ws.cell(row=tot_row, column=1, value="Total").font = bold_font
            ws.cell(row=tot_row, column=1).alignment = Alignment(horizontal="left")
            ws.cell(row=tot_row, column=1).border = total_border
            ws.cell(row=tot_row, column=1).fill = total_fill
            
            for col_num in range(2, len(data_df.columns) + 1):
                col_name = data_df.columns[col_num - 1]
                cell = ws.cell(row=tot_row, column=col_num)
                cell.border = total_border
                cell.fill = total_fill
                
                if any(x in col_name for x in ('val', 'igst', 'cgst', 'sgst', 'diff')):
                    col_letter = get_column_letter(col_num)
                    cell.value = f"=SUM({col_letter}2:{col_letter}{tot_row-1})"
                    cell.number_format = '#,##0.00'
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal="right")
                    
        # Autofit Columns
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                # Limit width calc for long law remarks
                if len(val_str) > 50:
                    val_str = val_str[:50]
                max_len = max(max_len, len(val_str))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 2. Write Sheets
    if not df_reco.empty:
        # SINGLE CONSOLIDATED RECONCILIATION SHEET
        # Reorder columns to make it extremely clear (Status, Remarks, Law Remarks first)
        ordered_cols = [
            'reco_status', 'itc_action', 'remarks', 'gst_law_remark',
            'books_gstin', 'books_supplier_name', 'books_doc_num', 'books_doc_date', 'books_doc_type',
            'books_taxable_val', 'books_igst', 'books_cgst', 'books_sgst', 'books_cess', 'books_total_val',
            'books_pos', 'books_rchrg', 'books_section',
            'gstr2b_gstin', 'gstr2b_supplier_name', 'gstr2b_doc_num', 'gstr2b_doc_date', 'gstr2b_doc_type',
            'gstr2b_taxable_val', 'gstr2b_igst', 'gstr2b_cgst', 'gstr2b_sgst', 'gstr2b_cess', 'gstr2b_total_val',
            'gstr2b_pos', 'gstr2b_rchrg', 'gstr2b_itc_eligibility', 'gstr2b_filing_date', 'gstr2b_gstr3b_status',
            'gstr2b_section', 'gstr2b_rtn_period', 'gstr2b_source_file',
            'taxable_val_diff', 'igst_diff', 'cgst_diff', 'sgst_diff', 'days_diff'
        ]
        
        # Select and align columns present in dataframe
        avail_cols = [c for c in ordered_cols if c in df_reco.columns]
        df_reco_ordered = df_reco[avail_cols]
        
        write_sheet(wb, "Reconciliation Report", df_reco_ordered, is_reconciliation_report=True)
        
    if not df_supplier.empty:
        # Supplier Performance Summary Tab
        write_sheet(wb, "Supplier Summary", df_supplier, is_reconciliation_report=False)
        
    # Output byte buffer
    out_io = io.BytesIO()
    wb.save(out_io)
    out_io.seek(0)
    return out_io

# Main Application Streamlit UI
st.title("💼 GSTR-2B Reconciliation Dashboard")
st.markdown("Reconcile GSTR-2B auto-drafted statements against internal Purchase Registers under active GST Rules & Laws.")

# Sidebar Configuration Controls
with st.sidebar:
    st.header("⚙️ Reconciliation Settings")
    
    # Tolerances
    val_tol = st.slider("Taxable Value Tolerance (₹)", min_value=0.0, max_value=100.0, value=10.0, step=1.0, 
                        help="Allowable difference in taxable value between books and portal records.")
    date_tol = st.slider("Date Tolerance (Days)", min_value=0, max_value=30, value=7, step=1,
                         help="Allowable variance between Books invoice date and Portal filing date.")
    fuzzy_tol = st.slider("Fuzzy Match Sensitivity (%)", min_value=70, max_value=100, value=85, step=5,
                          help="Minimum invoice number similarity score to classify as fuzzy matched.")
    
    cn_convention = st.selectbox(
        "Credit Note Value Sign",
        options=["auto", "negative", "keep_original"],
        index=0,
        help="How to represent Credit Notes. 'auto' converts notes to negative so they reduce ITC values naturally."
    )
    
    st.markdown("---")
    st.header("📂 Upload Invoices")
    
    # GSTR-2B File upload (1GB upload limit enabled)
    gstr2b_files = st.file_uploader(
        "GSTR-2B JSON Files (Up to 1GB)", 
        type="json", 
        accept_multiple_files=True,
        help="Upload one or multiple GSTR-2B JSON statements. Supports huge size files."
    )
    
    # Books File upload
    books_file = st.file_uploader(
        "Purchase Register (Excel/CSV - Up to 1GB)", 
        type=["xlsx", "xls", "csv"], 
        accept_multiple_files=False,
        help="Upload your Purchase Register sheet."
    )
    
    # Sandbox Testing Button
    st.markdown("---")
    st.subheader("💡 Sandbox Testing")
    if st.button("Generate & Load Synthetic Data"):
        json_content, df_books_synth = get_synthetic_data()
        st.session_state['synth_json'] = json_content
        st.session_state['synth_books'] = df_books_synth
        st.session_state['loaded_synth'] = True
        st.success("Loaded GSTR-2B JSON and Purchase Register mock datasets featuring ISD, RCM, SEZ, Late, and Blocked ITC entries.")

# State initialization
if 'loaded_synth' not in st.session_state:
    st.session_state['loaded_synth'] = False

# Load data logic
gstr2b_df = pd.DataFrame()
books_df = pd.DataFrame()
sheet_names = []
selected_sheet = None

# If user loaded synthetic data
if st.session_state['loaded_synth']:
    gstr2b_df = parse_gstr2b_json(json.dumps(st.session_state['synth_json']), "sample_gstr2b.json")
    
    col_mapping = {
        'supplier_gstin': 'Supplier GSTIN',
        'supplier_name': 'Supplier Name',
        'doc_num': 'Invoice Number',
        'doc_date': 'Invoice Date',
        'taxable_val': 'Taxable Value',
        'igst': 'IGST',
        'cgst': 'CGST',
        'sgst': 'SGST',
        'doc_type': 'Voucher Type',
        'pos': 'POS',
        'rchrg': 'RCM'
    }
    
    books_df = parse_purchase_register(
        st.session_state['synth_books'], 
        col_mapping, 
        credit_note_convention=cn_convention
    )
    st.info("⚡ Active Source: Sandbox Synthetic Datasets (ISD, SEZ, RCM, unfiled G3B return cases active)")
    
else:
    # 1. Parse Uploaded GSTR-2B JSON
    if gstr2b_files:
        dfs = []
        for file in gstr2b_files:
            file_name = file.name
            try:
                # Read content
                content = file.read().decode('utf-8')
                df = parse_gstr2b_json(content, file_name)
                dfs.append(df)
            except Exception as e:
                st.error(f"Error parsing GSTR-2B file {file_name}: {str(e)}")
        if dfs:
            gstr2b_df = pd.concat(dfs, ignore_index=True)
            st.success(f"Successfully parsed {len(gstr2b_files)} GSTR-2B JSON file(s) ({len(gstr2b_df)} total records).")

    # 2. Parse Uploaded Purchase Register
    if books_file:
        if books_file.name.lower().endswith('.xls'):
            try:
                xl = pd.ExcelFile(books_file)
                sheet_names = xl.sheet_names
            except Exception as e:
                st.error(f"Error reading legacy Excel sheets: {str(e)}")
        elif books_file.name.lower().endswith(('.xlsx', '.xlsm')):
            try:
                wb = load_workbook(books_file, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            except Exception as e:
                st.error(f"Error reading sheets from Excel: {str(e)}")
                
        if len(sheet_names) > 1:
            selected_sheet = st.selectbox("Select Excel Worksheet", options=sheet_names)
            
        try:
            # Quick head load to fetch headers
            books_file.seek(0)
            if books_file.name.endswith('.csv'):
                header_df = pd.read_csv(books_file, nrows=0)
            elif books_file.name.lower().endswith('.xls') or books_file.size <= 10 * 1024 * 1024:
                header_df = pd.read_excel(books_file, sheet_name=selected_sheet or 0, nrows=0)
            else:
                # Large xlsx file: use openpyxl read_only for low memory
                wb = load_workbook(books_file, read_only=True)
                ws = wb[selected_sheet] if selected_sheet else wb.active
                rows = ws.iter_rows(values_only=True)
                headers_raw = next(rows)
                wb.close()
                header_df = pd.DataFrame(columns=[h for h in headers_raw if h is not None])
                
            cols = list(header_df.columns)
            detected_maps = auto_detect_columns(cols)
            
            st.markdown("### 🗺️ Column Mapping")
            st.markdown("Align Purchase Register column headers to target fields. Auto-detected selections are filled.")
            
            map_cols = st.columns(3)
            col_mapping = {}
            
            # Map required fields
            fields_meta = [
                ('supplier_gstin', 'Supplier GSTIN *', True),
                ('doc_num', 'Invoice/Document Number *', True),
                ('doc_date', 'Invoice Date *', True),
                ('taxable_val', 'Taxable Value *', True),
                ('igst', 'IGST Amount', False),
                ('cgst', 'CGST Amount', False),
                ('sgst', 'SGST Amount', False),
                ('cess', 'Cess Amount', False),
                ('total_val', 'Total Value', False),
                ('supplier_name', 'Supplier Name', False),
                ('doc_type', 'Document Type column', False),
                ('pos', 'Place of Supply (POS)', False),
                ('rchrg', 'Reverse Charge (RCM)', False)
            ]
            
            for idx, (field_id, label, is_required) in enumerate(fields_meta):
                col_place = map_cols[idx % 3]
                default_val = detected_maps.get(field_id)
                default_idx = cols.index(default_val) if default_val in cols else 0
                
                with col_place:
                    sel = st.selectbox(
                        label,
                        options=["[Not Selected]"] + cols if not is_required else cols,
                        index=default_idx if is_required else (default_idx + 1 if default_val in cols else 0),
                        key=f"col_map_{field_id}"
                    )
                    if sel != "[Not Selected]":
                        col_mapping[field_id] = sel
                        
            if st.button("Load Purchase Register"):
                try:
                    books_file.seek(0)
                    books_df = parse_purchase_register(
                        books_file, 
                        col_mapping, 
                        sheet_name=selected_sheet, 
                        credit_note_convention=cn_convention
                    )
                    st.session_state['books_df_parsed'] = books_df
                    st.success(f"Purchase Register loaded successfully! ({len(books_df)} documents parsed).")
                except Exception as e:
                    st.error(f"Error loading Purchase Register: {str(e)}")
        except Exception as e:
            st.error(f"Error reading file structure: {str(e)}")

# If books register is loaded in state
if 'books_df_parsed' in st.session_state and not books_file is None:
    books_df = st.session_state['books_df_parsed']

# Reconciliation Trigger
if not gstr2b_df.empty and not books_df.empty:
    st.markdown("---")
    st.subheader("🏁 Run Reconciliation Engine")
    
    if st.button("Run Reconciliation", type="primary", use_container_width=True):
        with st.spinner("Executing smart matching engine (analyzing exact match, tolerances, amendments, fuzzy logic, RCM, SEZ, and ISD)..."):
            try:
                df_reco = reconcile_data(
                    books_df, 
                    gstr2b_df, 
                    val_tolerance=val_tol, 
                    date_tolerance_days=date_tol, 
                    fuzzy_threshold=fuzzy_tol
                )
                df_supplier = generate_supplier_summary(df_reco)
                
                # Save in session state
                st.session_state['reco_results'] = df_reco
                st.session_state['supplier_results'] = df_supplier
                st.session_state['reco_executed'] = True
                st.success("Reconciliation complete!")
            except Exception as e:
                st.error(f"Failed during reconciliation process: {str(e)}")

# Display Results
if 'reco_executed' in st.session_state and st.session_state['reco_executed']:
    df_reco = st.session_state['reco_results']
    df_supplier = st.session_state['supplier_results']
    
    if df_reco.empty:
        st.warning("Reconciliation finished but produced no output data.")
    else:
        # Totals in Books
        books_tax_total = df_reco['books_taxable_val'].sum()
        books_itc_total = df_reco['books_igst'].sum() + df_reco['books_cgst'].sum() + df_reco['books_sgst'].sum()
        
        # Totals GSTR-2B
        g2b_tax_total = df_reco['gstr2b_taxable_val'].sum()
        g2b_itc_total = df_reco['gstr2b_igst'].sum() + df_reco['gstr2b_cgst'].sum() + df_reco['gstr2b_sgst'].sum()
        
        # Matches count
        total_rows = len(df_reco)
        matched_rows = df_reco[df_reco['reco_status'].isin(['Matched', 'Fuzzy Match', 'Matched (Amended)'])].shape[0]
        mismatch_rows = df_reco[df_reco['reco_status'].str.contains('Mismatch', na=False)].shape[0]
        only_books_rows = df_reco[df_reco['reco_status'] == 'Only in Books'].shape[0]
        only_2b_rows = df_reco[df_reco['reco_status'] == 'Only in GSTR-2B'].shape[0]
        
        # Eligible ITC Claimable
        claimable_itc = df_reco[df_reco['itc_action'].str.contains('Claimable|Claim', na=False)]
        claimable_itc_total = (
            claimable_itc['gstr2b_igst'].sum() + 
            claimable_itc['gstr2b_cgst'].sum() + 
            claimable_itc['gstr2b_sgst'].sum()
        )
        
        summary_stats = {
            'books_count': int(df_reco['books_doc_num'].dropna().nunique()),
            'gstr2b_count': int(df_reco['gstr2b_doc_num'].dropna().nunique()),
            'books_taxable_total': float(books_tax_total),
            'books_igst_total': float(df_reco['books_igst'].sum()),
            'books_cgst_total': float(df_reco['books_cgst'].sum()),
            'books_sgst_total': float(df_reco['books_sgst'].sum()),
            'gstr2b_taxable_total': float(g2b_tax_total),
            'gstr2b_igst_total': float(df_reco['gstr2b_igst'].sum()),
            'gstr2b_cgst_total': float(df_reco['gstr2b_cgst'].sum()),
            'gstr2b_sgst_total': float(df_reco['gstr2b_sgst'].sum()),
            'taxable_variance': float(books_tax_total - g2b_tax_total),
            'igst_variance': float(df_reco['books_igst'].sum() - df_reco['gstr2b_igst'].sum()),
            'cgst_variance': float(df_reco['books_cgst'].sum() - df_reco['gstr2b_cgst'].sum()),
            'sgst_variance': float(df_reco['books_sgst'].sum() - df_reco['gstr2b_sgst'].sum()),
            'claimable_itc_total': float(claimable_itc_total),
            'gstr2b_itc_total': float(g2b_itc_total),
        }

        # Generate styled Excel single-sheet output
        excel_bytes = export_reco_to_excel(df_reco, df_supplier, summary_stats)
        
        st.markdown("### 📊 Reconciliation Summary Report")
        
        kpi_cols = st.columns(4)
        with kpi_cols[0]:
            st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-title">Matched Accuracy Rate</div>
                    <div class="kpi-value">{round((matched_rows/total_rows)*100, 1) if total_rows > 0 else 0.0}%</div>
                    <div class="kpi-subtitle">{matched_rows} matched / {total_rows} total rows</div>
                </div>
            """, unsafe_allow_html=True)
            
        with kpi_cols[1]:
            st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-title">Total Claimable ITC</div>
                    <div class="kpi-value">₹ {claimable_itc_total:,.2f}</div>
                    <div class="kpi-subtitle">Eligible match from GSTR-2B</div>
                </div>
            """, unsafe_allow_html=True)
            
        with kpi_cols[2]:
            st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-title">Books vs GSTR-2B Gap</div>
                    <div class="kpi-value" style="color: {'#10B981' if abs(books_itc_total - g2b_itc_total) < 10 else '#F59E0B'}">₹ {(books_itc_total - g2b_itc_total):,.2f}</div>
                    <div class="kpi-subtitle">Books ITC: ₹{books_itc_total:,.2f} | 2B ITC: ₹{g2b_itc_total:,.2f}</div>
                </div>
            """, unsafe_allow_html=True)
            
        with kpi_cols[3]:
            st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-title">Actionable Mismatches</div>
                    <div class="kpi-value" style="color: {'#F43F5E' if mismatch_rows > 0 else '#F8FAFC'}">{mismatch_rows}</div>
                    <div class="kpi-subtitle">Value or Date variance errors</div>
                </div>
            """, unsafe_allow_html=True)

        # Excel Download Button (Single consolidated Reconciliation sheet + summary tabs)
        st.download_button(
            label="📥 Download Excel Reconciliation Report (Single Reconciled sheet with comments)",
            data=excel_bytes,
            file_name="GSTR2B_Reconciliation_Report_Consolidated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        # Visualizations (Plotly)
        st.markdown("---")
        vis_cols = st.columns(2)
        
        with vis_cols[0]:
            st.subheader("Matching Status Breakdown")
            status_counts = df_reco['reco_status'].value_counts().reset_index()
            status_counts.columns = ['Status', 'Count']
            
            fig_pie = px.pie(
                status_counts, 
                values='Count', 
                names='Status',
                color='Status',
                color_discrete_map={
                    'Matched': '#10B981',
                    'Fuzzy Match': '#06B6D4',
                    'Matched (Amended)': '#34D399',
                    'Value Mismatch': '#F59E0B',
                    'Date Mismatch': '#F97316',
                    'Date & Value Mismatch': '#EF4444',
                    'Value Mismatch (Amended)': '#D97706',
                    'Only in Books': '#F43F5E',
                    'Only in GSTR-2B': '#6366F1'
                },
                hole=0.4
            )
            fig_pie.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=300, showlegend=True)
            st.plotly_chart(fig_pie, use_container_width=True)
            
        with vis_cols[1]:
            st.subheader("Books vs GSTR-2B ITC Comparison")
            fig_bar = go.Figure(data=[
                go.Bar(name='Purchase Register (Books)', x=['IGST', 'CGST', 'SGST'], y=[
                    df_reco['books_igst'].sum(),
                    df_reco['books_cgst'].sum(),
                    df_reco['books_sgst'].sum()
                ], marker_color='#F43F5E'),
                go.Bar(name='GSTR-2B Portal', x=['IGST', 'CGST', 'SGST'], y=[
                    df_reco['gstr2b_igst'].sum(),
                    df_reco['gstr2b_cgst'].sum(),
                    df_reco['gstr2b_sgst'].sum()
                ], marker_color='#6366F1')
            ])
            fig_bar.update_layout(barmode='group', margin=dict(t=10, b=10, l=10, r=10), height=300)
            st.plotly_chart(fig_bar, use_container_width=True)

        # Tabbed Data Explorer
        st.markdown("---")
        st.subheader("🔍 Interactive Data Explorer")
        
        # Display Column selection
        show_compliance_cols = st.checkbox("Show Compliance Columns (POS, RCM, Section, Filing Dates, GST Law)", value=True)
        
        # Filters
        filter_cols = st.columns(3)
        with filter_cols[0]:
            gstin_filter = st.selectbox(
                "Filter by Supplier GSTIN",
                options=["All"] + sorted(list(df_reco['books_gstin'].fillna(df_reco['gstr2b_gstin']).dropna().unique()))
            )
        with filter_cols[1]:
            status_filter = st.selectbox(
                "Filter by Matching Status",
                options=["All", "Matched / Fuzzy Match", "Discrepancy / Mismatch", "Only in Books", "Only in GSTR-2B"]
            )
        with filter_cols[2]:
            search_inv = st.text_input("Search Invoice Number", value="", placeholder="Enter invoice number...")
            
        # Apply filters
        df_filtered = df_reco.copy()
        if gstin_filter != "All":
            df_filtered = df_filtered[(df_filtered['books_gstin'] == gstin_filter) | (df_filtered['gstr2b_gstin'] == gstin_filter)]
            
        if status_filter == "Matched / Fuzzy Match":
            df_filtered = df_filtered[df_filtered['reco_status'].isin(['Matched', 'Fuzzy Match', 'Matched (Amended)'])]
        elif status_filter == "Discrepancy / Mismatch":
            df_filtered = df_filtered[df_filtered['reco_status'].str.contains('Mismatch', na=False)]
        elif status_filter == "Only in Books":
            df_filtered = df_filtered[df_filtered['reco_status'] == 'Only in Books']
        elif status_filter == "Only in GSTR-2B":
            df_filtered = df_filtered[df_filtered['reco_status'] == 'Only in GSTR-2B']
            
        if search_inv:
            clean_search = clean_invoice_number(search_inv)
            df_filtered = df_filtered[
                df_filtered['books_doc_num'].astype(str).str.contains(search_inv, case=False, na=False) |
                df_filtered['gstr2b_doc_num'].astype(str).str.contains(search_inv, case=False, na=False) |
                df_filtered['books_doc_num'].apply(clean_invoice_number).str.contains(clean_search, na=False) |
                df_filtered['gstr2b_doc_num'].apply(clean_invoice_number).str.contains(clean_search, na=False)
            ]

        explorer_tabs = st.tabs([
            "Reconciliation Report (Single Sheet view)", 
            "Supplier Summary"
        ])
        
        # Color coding rows
        def highlight_status(row):
            status = row['reco_status']
            if status in ('Matched', 'Fuzzy Match', 'Matched (Amended)'):
                return ['background-color: rgba(16, 185, 129, 0.08)'] * len(row)
            elif 'Mismatch' in status:
                return ['background-color: rgba(245, 158, 11, 0.08)'] * len(row)
            elif status == 'Only in Books':
                return ['background-color: rgba(244, 63, 94, 0.08)'] * len(row)
            elif status == 'Only in GSTR-2B':
                return ['background-color: rgba(99, 102, 241, 0.08)'] * len(row)
            return [''] * len(row)

        with explorer_tabs[0]:
            st.markdown(f"Showing **{len(df_filtered)}** reconciled documents. Remarks details show value and date discrepancy comments.")
            if not df_filtered.empty:
                df_disp = df_filtered.copy()
                df_disp['books_doc_date'] = df_disp['books_doc_date'].dt.strftime('%d-%m-%Y')
                df_disp['gstr2b_doc_date'] = df_disp['gstr2b_doc_date'].dt.strftime('%d-%m-%Y')
                df_disp['gstr2b_filing_date'] = df_disp['gstr2b_filing_date'].dt.strftime('%d-%m-%Y')
                
                # Standard display columns
                col_order_list = [
                    'reco_status', 'itc_action', 'remarks'
                ]
                
                if show_compliance_cols:
                    col_order_list.append('gst_law_remark')
                    
                col_order_list.extend([
                    'books_gstin', 'books_doc_num', 'books_doc_date', 'books_taxable_val', 'books_igst', 'books_cgst', 'books_sgst'
                ])
                
                if show_compliance_cols:
                    col_order_list.extend(['books_pos', 'books_rchrg'])
                    
                col_order_list.extend([
                    'gstr2b_gstin', 'gstr2b_doc_num', 'gstr2b_doc_date', 'gstr2b_taxable_val', 'gstr2b_igst', 'gstr2b_cgst', 'gstr2b_sgst'
                ])
                
                if show_compliance_cols:
                    col_order_list.extend(['gstr2b_pos', 'gstr2b_rchrg', 'gstr2b_itc_eligibility', 'gstr2b_filing_date', 'gstr2b_gstr3b_status', 'gstr2b_section'])
                    
                col_order_list.extend([
                    'taxable_val_diff', 'igst_diff'
                ])
                
                st.dataframe(
                    df_disp.style.apply(highlight_status, axis=1),
                    column_order=col_order_list,
                    use_container_width=True,
                    height=500
                )
            else:
                st.info("No records matching current filters.")
                
        with explorer_tabs[1]:
            st.markdown("Grouped supplier-wise summary of reconciliation counts and tax values.")
            if not df_supplier.empty:
                df_supp_filtered = df_supplier.copy()
                if gstin_filter != "All":
                    df_supp_filtered = df_supp_filtered[df_supp_filtered['supplier_gstin'] == gstin_filter]
                
                st.dataframe(
                    df_supp_filtered,
                    column_order=[
                        'supplier_gstin', 'supplier_name', 'books_invoice_count', 'gstr2b_invoice_count',
                        'books_taxable_val', 'books_total_itc', 'gstr2b_taxable_val', 'gstr2b_total_itc',
                        'taxable_val_diff', 'itc_diff', 'match_rate_pct'
                    ],
                    use_container_width=True,
                    height=450
                )
            else:
                st.info("No supplier summaries available.")
else:
    st.info("👋 Welcome! Please upload GSTR-2B JSON files and your Purchase Register in the sidebar, or click the **Generate & Load Synthetic Data** button to explore the dashboard immediately.")
