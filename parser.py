import json
import re
import pandas as pd
import numpy as np
import gc
from openpyxl import load_workbook

def get_json_section(data, key):
    """
    Search recursively or through common GSTR-2B JSON wrappers for a specific key.
    """
    if not isinstance(data, dict):
        return []
    
    # Direct match
    if key in data:
        return data[key]
    
    # Check standard 'data' wrappers
    if 'data' in data:
        inner = data['data']
        if isinstance(inner, dict):
            if key in inner:
                return inner[key]
            if 'data' in inner:
                double_inner = inner['data']
                if isinstance(double_inner, dict) and key in double_inner:
                    return double_inner[key]
                
    # Exhaustive search for the key (as a fallback)
    for k, v in data.items():
        if k == key and isinstance(v, list):
            return v
        if isinstance(v, dict):
            res = get_json_section(v, key)
            if res:
                return res
            
    return []

def clean_invoice_number(inv_no):
    """
    Cleans invoice number for robust matching:
    - Standardizes to uppercase.
    - Strips whitespace.
    - Removes non-alphanumeric characters.
    - Strips leading zeros.
    """
    if pd.isna(inv_no) or inv_no is None:
        return ""
    inv_str = str(inv_no).strip().upper()
    # Remove all non-alphanumeric characters (e.g., slash, dash, spaces)
    cleaned = re.sub(r'[^A-Z0-9]', '', inv_str)
    # Strip leading zeros
    cleaned = cleaned.lstrip('0')
    return cleaned

def parse_date(date_val):
    """
    Parses various date formats to pd.Timestamp.
    """
    if pd.isna(date_val) or date_val is None:
        return pd.NaT
    if isinstance(date_val, pd.Timestamp):
        return date_val
    if isinstance(date_val, (int, float)):
        # Excel numeric date handling
        try:
            return pd.to_datetime(date_val, unit='D', origin='1899-12-30')
        except:
            return pd.NaT
            
    date_str = str(date_val).strip()
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%b-%Y', '%d %b %Y'):
        try:
            return pd.to_datetime(date_str, format=fmt)
        except ValueError:
            continue
    try:
        return pd.to_datetime(date_str)
    except:
        return pd.NaT

def parse_gstr2b_json(json_content_or_path, file_name="GSTR2B.json"):
    """
    Parses a GSTR-2B JSON file and flattens it into a pandas DataFrame.
    Supports B2B, B2BA, CDNR, CDNRA, ISD, ISDA, IMPG, and IMPGSEZ sections.
    """
    if isinstance(json_content_or_path, str):
        try:
            with open(json_content_or_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            # Try load as raw string if path fails
            data = json.loads(json_content_or_path)
    else:
        # File-like object
        data = json.load(json_content_or_path)

    # Attempt to extract return period and recipient GSTIN
    rtn_period = data.get('rtnprd') or data.get('fp')
    recipient_gstin = data.get('gstin') or data.get('rcptgstin')
    
    if not rtn_period and 'data' in data and isinstance(data['data'], dict):
        rtn_period = data['data'].get('rtnprd') or data['data'].get('fp')
        recipient_gstin = data['data'].get('gstin') or data['data'].get('rcptgstin')

    documents = []

    # Helper function to extract document list from standard formats
    def get_doc_list(parent_obj):
        if 'doclist' in parent_obj:
            return parent_obj['doclist']
        if 'inv' in parent_obj:
            return parent_obj['inv']
        if 'nt' in parent_obj:
            return parent_obj['nt']
        if 'boe' in parent_obj:
            return parent_obj['boe']
        return []

    # 1. Parse B2B Invoices
    b2b_list = get_json_section(data, 'b2b')
    for supplier in b2b_list:
        ctin = supplier.get('ctin', '').strip().upper()
        cname = supplier.get('lglNm') or supplier.get('trdNm') or ""
        doc_list = get_doc_list(supplier)
        
        for inv in doc_list:
            inum = inv.get('inum', '').strip()
            idt = inv.get('idt') or inv.get('dt')
            val = float(inv.get('val', 0.0))
            itcelg = inv.get('itcelg', 'Y').strip().upper()
            pos = inv.get('pos') or ""
            rchrg = inv.get('rchrg') or inv.get('rc') or 'N'
            filing_date = inv.get('flddt') or inv.get('fld_dt') or inv.get('filing_date') or ""
            gstr3b_status = inv.get('g3bfil') or inv.get('g3bFilingStatus') or inv.get('g3bStatus') or 'N'
            
            items = inv.get('items', [])
            txval = sum(float(item.get('txval', 0.0)) for item in items)
            igst = sum(float(item.get('igst', 0.0)) for item in items)
            cgst = sum(float(item.get('cgst', 0.0)) for item in items)
            sgst = sum(float(item.get('sgst', 0.0)) for item in items)
            cess = sum(float(item.get('cess', 0.0)) for item in items)
            
            documents.append({
                'supplier_gstin': ctin,
                'supplier_name': cname,
                'doc_num': inum,
                'clean_doc_num': clean_invoice_number(inum),
                'doc_date': parse_date(idt),
                'doc_type': 'INV',
                'taxable_val': round(txval, 2),
                'igst': round(igst, 2),
                'cgst': round(cgst, 2),
                'sgst': round(sgst, 2),
                'cess': round(cess, 2),
                'total_val': round(val, 2),
                'pos': pos,
                'rchrg': 'Yes' if rchrg in ('Y', 'YES') else 'No',
                'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
                'filing_date': parse_date(filing_date),
                'gstr3b_status': 'Yes' if gstr3b_status in ('Y', 'YES', 'FILING_STATUS_FILED') else 'No',
                'section': 'B2B Invoices',
                'is_amended': False,
                'original_doc_num': None,
                'original_doc_date': None,
                'rtn_period': rtn_period,
                'source': 'GSTR-2B',
                'source_file': file_name
            })

    # 2. Parse B2BA (Amended B2B Invoices)
    b2ba_list = get_json_section(data, 'b2ba')
    for supplier in b2ba_list:
        ctin = supplier.get('ctin', '').strip().upper()
        cname = supplier.get('lglNm') or supplier.get('trdNm') or ""
        doc_list = get_doc_list(supplier)
        
        for inv in doc_list:
            inum = inv.get('inum', '').strip()
            idt = inv.get('idt') or inv.get('dt')
            oinum = inv.get('oinum', '').strip()
            oidt = inv.get('oidt')
            val = float(inv.get('val', 0.0))
            itcelg = inv.get('itcelg', 'Y').strip().upper()
            pos = inv.get('pos') or ""
            rchrg = inv.get('rchrg') or inv.get('rc') or 'N'
            filing_date = inv.get('flddt') or inv.get('fld_dt') or inv.get('filing_date') or ""
            gstr3b_status = inv.get('g3bfil') or inv.get('g3bFilingStatus') or inv.get('g3bStatus') or 'N'
            
            items = inv.get('items', [])
            txval = sum(float(item.get('txval', 0.0)) for item in items)
            igst = sum(float(item.get('igst', 0.0)) for item in items)
            cgst = sum(float(item.get('cgst', 0.0)) for item in items)
            sgst = sum(float(item.get('sgst', 0.0)) for item in items)
            cess = sum(float(item.get('cess', 0.0)) for item in items)
            
            documents.append({
                'supplier_gstin': ctin,
                'supplier_name': cname,
                'doc_num': inum,
                'clean_doc_num': clean_invoice_number(inum),
                'doc_date': parse_date(idt),
                'doc_type': 'INV',
                'taxable_val': round(txval, 2),
                'igst': round(igst, 2),
                'cgst': round(cgst, 2),
                'sgst': round(sgst, 2),
                'cess': round(cess, 2),
                'total_val': round(val, 2),
                'pos': pos,
                'rchrg': 'Yes' if rchrg in ('Y', 'YES') else 'No',
                'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
                'filing_date': parse_date(filing_date),
                'gstr3b_status': 'Yes' if gstr3b_status in ('Y', 'YES', 'FILING_STATUS_FILED') else 'No',
                'section': 'B2B Amendments',
                'is_amended': True,
                'original_doc_num': oinum,
                'original_doc_date': parse_date(oidt),
                'rtn_period': rtn_period,
                'source': 'GSTR-2B',
                'source_file': file_name
            })

    # 3. Parse CDNR (Credit / Debit Notes)
    cdnr_list = get_json_section(data, 'cdnr')
    for supplier in cdnr_list:
        ctin = supplier.get('ctin', '').strip().upper()
        cname = supplier.get('lglNm') or supplier.get('trdNm') or ""
        doc_list = get_doc_list(supplier)
        
        for nt in doc_list:
            nt_num = (nt.get('nt_num') or nt.get('ntnum') or '').strip()
            nt_dt = nt.get('nt_dt') or nt.get('ntdt') or nt.get('dt')
            nt_ty = nt.get('nt_ty', '').strip().upper() # C = Credit Note, D = Debit Note
            val = float(nt.get('val', 0.0))
            itcelg = nt.get('itcelg', 'Y').strip().upper()
            pos = nt.get('pos') or ""
            rchrg = nt.get('rchrg') or nt.get('rc') or 'N'
            filing_date = nt.get('flddt') or nt.get('fld_dt') or nt.get('filing_date') or ""
            gstr3b_status = nt.get('g3bfil') or nt.get('g3bFilingStatus') or nt.get('g3bStatus') or 'N'
            
            items = nt.get('items', [])
            txval = sum(float(item.get('txval', 0.0)) for item in items)
            igst = sum(float(item.get('igst', 0.0)) for item in items)
            cgst = sum(float(item.get('cgst', 0.0)) for item in items)
            sgst = sum(float(item.get('sgst', 0.0)) for item in items)
            cess = sum(float(item.get('cess', 0.0)) for item in items)
            
            # Credit notes represent negative values (ITC reduction)
            sign = -1.0 if nt_ty == 'C' else 1.0
            doc_type = 'CRN' if nt_ty == 'C' else 'DBN'
            
            documents.append({
                'supplier_gstin': ctin,
                'supplier_name': cname,
                'doc_num': nt_num,
                'clean_doc_num': clean_invoice_number(nt_num),
                'doc_date': parse_date(nt_dt),
                'doc_type': doc_type,
                'taxable_val': round(txval * sign, 2),
                'igst': round(igst * sign, 2),
                'cgst': round(cgst * sign, 2),
                'sgst': round(sgst * sign, 2),
                'cess': round(cess * sign, 2),
                'total_val': round(val * sign, 2),
                'pos': pos,
                'rchrg': 'Yes' if rchrg in ('Y', 'YES') else 'No',
                'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
                'filing_date': parse_date(filing_date),
                'gstr3b_status': 'Yes' if gstr3b_status in ('Y', 'YES', 'FILING_STATUS_FILED') else 'No',
                'section': 'Credit/Debit Notes',
                'is_amended': False,
                'original_doc_num': None,
                'original_doc_date': None,
                'rtn_period': rtn_period,
                'source': 'GSTR-2B',
                'source_file': file_name
            })

    # 4. Parse CDNRA (Amended Credit / Debit Notes)
    cdnra_list = get_json_section(data, 'cdnra')
    for supplier in cdnra_list:
        ctin = supplier.get('ctin', '').strip().upper()
        cname = supplier.get('lglNm') or supplier.get('trdNm') or ""
        doc_list = get_doc_list(supplier)
        
        for nt in doc_list:
            nt_num = (nt.get('nt_num') or nt.get('ntnum') or '').strip()
            nt_dt = nt.get('nt_dt') or nt.get('ntdt') or nt.get('dt')
            ont_num = (nt.get('ont_num') or nt.get('ontnum') or '').strip()
            ont_dt = nt.get('ont_dt') or nt.get('ontdt')
            nt_ty = nt.get('nt_ty', '').strip().upper()
            val = float(nt.get('val', 0.0))
            itcelg = nt.get('itcelg', 'Y').strip().upper()
            pos = nt.get('pos') or ""
            rchrg = nt.get('rchrg') or nt.get('rc') or 'N'
            filing_date = nt.get('flddt') or nt.get('fld_dt') or nt.get('filing_date') or ""
            gstr3b_status = nt.get('g3bfil') or nt.get('g3bFilingStatus') or nt.get('g3bStatus') or 'N'
            
            items = nt.get('items', [])
            txval = sum(float(item.get('txval', 0.0)) for item in items)
            igst = sum(float(item.get('igst', 0.0)) for item in items)
            cgst = sum(float(item.get('cgst', 0.0)) for item in items)
            sgst = sum(float(item.get('sgst', 0.0)) for item in items)
            cess = sum(float(item.get('cess', 0.0)) for item in items)
            
            sign = -1.0 if nt_ty == 'C' else 1.0
            doc_type = 'CRN' if nt_ty == 'C' else 'DBN'
            
            documents.append({
                'supplier_gstin': ctin,
                'supplier_name': cname,
                'doc_num': nt_num,
                'clean_doc_num': clean_invoice_number(nt_num),
                'doc_date': parse_date(nt_dt),
                'doc_type': doc_type,
                'taxable_val': round(txval * sign, 2),
                'igst': round(igst * sign, 2),
                'cgst': round(cgst * sign, 2),
                'sgst': round(sgst * sign, 2),
                'cess': round(cess * sign, 2),
                'total_val': round(val * sign, 2),
                'pos': pos,
                'rchrg': 'Yes' if rchrg in ('Y', 'YES') else 'No',
                'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
                'filing_date': parse_date(filing_date),
                'gstr3b_status': 'Yes' if gstr3b_status in ('Y', 'YES', 'FILING_STATUS_FILED') else 'No',
                'section': 'Credit/Debit Notes Amendments',
                'is_amended': True,
                'original_doc_num': ont_num,
                'original_doc_date': parse_date(ont_dt),
                'rtn_period': rtn_period,
                'source': 'GSTR-2B',
                'source_file': file_name
            })

    # 5. Parse ISD Invoices
    isd_list = get_json_section(data, 'isd')
    for supplier in isd_list:
        ctin = supplier.get('ctin', '').strip().upper()
        cname = supplier.get('lglNm') or supplier.get('trdNm') or ""
        doc_list = get_doc_list(supplier)
        
        for inv in doc_list:
            docnum = (inv.get('docnum') or inv.get('doc_num') or inv.get('inum') or '').strip()
            docdt = inv.get('docdt') or inv.get('doc_dt') or inv.get('idt')
            val = float(inv.get('val', 0.0))
            itcelg = inv.get('itcelg', 'Y').strip().upper()
            pos = inv.get('pos') or ""
            filing_date = inv.get('flddt') or inv.get('fld_dt') or inv.get('filing_date') or ""
            gstr3b_status = inv.get('g3bfil') or inv.get('g3bFilingStatus') or inv.get('g3bStatus') or 'Y'
            
            # ISD distribution values may be split rate-wise, sum them up
            items = inv.get('items', [])
            txval = sum(float(item.get('txval', 0.0)) for item in items)
            igst = sum(float(item.get('igst', 0.0)) for item in items)
            cgst = sum(float(item.get('cgst', 0.0)) for item in items)
            sgst = sum(float(item.get('sgst', 0.0)) for item in items)
            cess = sum(float(item.get('cess', 0.0)) for item in items)
            
            # Fallback if items are missing
            if not items:
                txval = float(inv.get('txval') or val)
                igst = float(inv.get('igst') or 0.0)
                cgst = float(inv.get('cgst') or 0.0)
                sgst = float(inv.get('sgst') or 0.0)
                cess = float(inv.get('cess') or 0.0)

            documents.append({
                'supplier_gstin': ctin,
                'supplier_name': cname,
                'doc_num': docnum,
                'clean_doc_num': clean_invoice_number(docnum),
                'doc_date': parse_date(docdt),
                'doc_type': 'INV',
                'taxable_val': round(txval, 2),
                'igst': round(igst, 2),
                'cgst': round(cgst, 2),
                'sgst': round(sgst, 2),
                'cess': round(cess, 2),
                'total_val': round(val, 2),
                'pos': pos,
                'rchrg': 'No', # ISD distribute input credits, no reverse charge
                'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
                'filing_date': parse_date(filing_date),
                'gstr3b_status': 'Yes' if gstr3b_status in ('Y', 'YES') else 'No',
                'section': 'ISD Invoices',
                'is_amended': False,
                'original_doc_num': None,
                'original_doc_date': None,
                'rtn_period': rtn_period,
                'source': 'GSTR-2B',
                'source_file': file_name
            })

    # 6. Parse ISDA Invoices (ISD Amendments)
    isda_list = get_json_section(data, 'isda')
    for supplier in isda_list:
        ctin = supplier.get('ctin', '').strip().upper()
        cname = supplier.get('lglNm') or supplier.get('trdNm') or ""
        doc_list = get_doc_list(supplier)
        
        for inv in doc_list:
            docnum = (inv.get('docnum') or inv.get('doc_num') or inv.get('inum') or '').strip()
            docdt = inv.get('docdt') or inv.get('doc_dt') or inv.get('idt')
            odocnum = (inv.get('odocnum') or inv.get('odoc_num') or inv.get('oinum') or '').strip()
            odocdt = inv.get('odocdt') or inv.get('odoc_dt') or inv.get('oidt')
            val = float(inv.get('val', 0.0))
            itcelg = inv.get('itcelg', 'Y').strip().upper()
            pos = inv.get('pos') or ""
            filing_date = inv.get('flddt') or inv.get('fld_dt') or inv.get('filing_date') or ""
            gstr3b_status = inv.get('g3bfil') or inv.get('g3bFilingStatus') or inv.get('g3bStatus') or 'Y'
            
            items = inv.get('items', [])
            txval = sum(float(item.get('txval', 0.0)) for item in items)
            igst = sum(float(item.get('igst', 0.0)) for item in items)
            cgst = sum(float(item.get('cgst', 0.0)) for item in items)
            sgst = sum(float(item.get('sgst', 0.0)) for item in items)
            cess = sum(float(item.get('cess', 0.0)) for item in items)
            
            if not items:
                txval = float(inv.get('txval') or val)
                igst = float(inv.get('igst') or 0.0)
                cgst = float(inv.get('cgst') or 0.0)
                sgst = float(inv.get('sgst') or 0.0)
                cess = float(inv.get('cess') or 0.0)

            documents.append({
                'supplier_gstin': ctin,
                'supplier_name': cname,
                'doc_num': docnum,
                'clean_doc_num': clean_invoice_number(docnum),
                'doc_date': parse_date(docdt),
                'doc_type': 'INV',
                'taxable_val': round(txval, 2),
                'igst': round(igst, 2),
                'cgst': round(cgst, 2),
                'sgst': round(sgst, 2),
                'cess': round(cess, 2),
                'total_val': round(val, 2),
                'pos': pos,
                'rchrg': 'No',
                'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
                'filing_date': parse_date(filing_date),
                'gstr3b_status': 'Yes' if gstr3b_status in ('Y', 'YES') else 'No',
                'section': 'ISD Amendments',
                'is_amended': True,
                'original_doc_num': odocnum,
                'original_doc_date': parse_date(odocdt),
                'rtn_period': rtn_period,
                'source': 'GSTR-2B',
                'source_file': file_name
            })

    # 7. Parse IMPG (Import of Goods)
    impg_list = get_json_section(data, 'impg')
    for boe in impg_list:
        boe_num = str(boe.get('boe_num') or boe.get('boenum') or boe.get('boenm') or '').strip()
        boe_dt = boe.get('boe_dt') or boe.get('boedt')
        val = float(boe.get('boe_val') or boe.get('val', 0.0))
        txval = float(boe.get('txval', 0.0))
        igst = float(boe.get('igst', 0.0))
        cess = float(boe.get('cess', 0.0))
        itcelg = boe.get('itcelg', 'Y').strip().upper()
        port_cd = boe.get('port_cd') or boe.get('port_code') or 'CUSTOMS'
        
        documents.append({
            'supplier_gstin': 'IMPORT',
            'supplier_name': f'Import from Port: {port_cd}',
            'doc_num': boe_num,
            'clean_doc_num': clean_invoice_number(boe_num),
            'doc_date': parse_date(boe_dt),
            'doc_type': 'BOE',
            'taxable_val': round(txval, 2),
            'igst': round(igst, 2),
            'cgst': 0.0,
            'sgst': 0.0,
            'cess': round(cess, 2),
            'total_val': round(val, 2),
            'pos': '97', # Outside India State Code
            'rchrg': 'No',
            'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
            'filing_date': pd.NaT,
            'gstr3b_status': 'Yes',
            'section': 'Import of Goods',
            'is_amended': False,
            'original_doc_num': None,
            'original_doc_date': None,
            'rtn_period': rtn_period,
            'source': 'GSTR-2B',
            'source_file': file_name
        })

    # 8. Parse IMPGSEZ (Import from SEZ units)
    impgsez_list = get_json_section(data, 'impgsez')
    for boe in impgsez_list:
        boe_num = str(boe.get('boe_num') or boe.get('boenum') or boe.get('boenm') or '').strip()
        boe_dt = boe.get('boe_dt') or boe.get('boedt')
        val = float(boe.get('boe_val') or boe.get('val', 0.0))
        txval = float(boe.get('txval', 0.0))
        igst = float(boe.get('igst', 0.0))
        cess = float(boe.get('cess', 0.0))
        itcelg = boe.get('itcelg', 'Y').strip().upper()
        
        # SEZ imports usually have the actual SEZ supplier's GSTIN
        ctin = (boe.get('ctin') or boe.get('gstin') or 'SEZ-IMPORT').strip().upper()
        cname = boe.get('lglNm') or boe.get('trdNm') or "SEZ Supplier"
        pos = boe.get('pos') or ""

        documents.append({
            'supplier_gstin': ctin,
            'supplier_name': cname,
            'doc_num': boe_num,
            'clean_doc_num': clean_invoice_number(boe_num),
            'doc_date': parse_date(boe_dt),
            'doc_type': 'BOE',
            'taxable_val': round(txval, 2),
            'igst': round(igst, 2),
            'cgst': 0.0,
            'sgst': 0.0,
            'cess': round(cess, 2),
            'total_val': round(val, 2),
            'pos': pos,
            'rchrg': 'No',
            'itc_eligibility': 'Eligible' if itcelg in ('Y', 'YES') else 'Ineligible',
            'filing_date': pd.NaT,
            'gstr3b_status': 'Yes',
            'section': 'Import from SEZ',
            'is_amended': False,
            'original_doc_num': None,
            'original_doc_date': None,
            'rtn_period': rtn_period,
            'source': 'GSTR-2B',
            'source_file': file_name
        })

    # Clean data & free memory
    del data
    gc.collect()

    # Create dataframe
    if not documents:
        return pd.DataFrame(columns=[
            'supplier_gstin', 'supplier_name', 'doc_num', 'clean_doc_num', 
            'doc_date', 'doc_type', 'taxable_val', 'igst', 'cgst', 'sgst', 
            'cess', 'total_val', 'pos', 'rchrg', 'itc_eligibility', 
            'filing_date', 'gstr3b_status', 'section', 'is_amended', 
            'original_doc_num', 'original_doc_date', 'rtn_period', 'source', 'source_file'
        ])
        
    df = pd.DataFrame(documents)
    return df

def parse_large_excel_streaming(file_path_or_buffer, sheet_name=None):
    """
    Streaming row-by-row Excel parser to read large spreadsheets (up to 1GB)
    with low memory overhead.
    """
    wb = load_workbook(file_path_or_buffer, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    rows = ws.iter_rows(values_only=True)
    
    # Read headers
    try:
        headers_raw = next(rows)
    except StopIteration:
        wb.close()
        return pd.DataFrame()
        
    headers = []
    # Fill in None or duplicate headers dynamically
    for idx, h in enumerate(headers_raw):
        if h is None:
            headers.append(f"Column_{idx}")
        else:
            h_str = str(h).strip()
            if h_str in headers:
                headers.append(f"{h_str}_{idx}")
            else:
                headers.append(h_str)

    data = []
    # Stream rows
    for r in rows:
        # Avoid loading extra empty rows at the end of Excel
        if all(val is None for val in r):
            continue
        # Zip row content to dict matching headers size
        row_dict = dict(zip(headers, r[:len(headers)]))
        data.append(row_dict)
        
    df_raw = pd.DataFrame(data)
    wb.close()
    
    # Force clean garbage collection
    del data
    gc.collect()
    
    return df_raw

def parse_purchase_register(file_path_or_buffer, col_mapping, sheet_name=None, credit_note_convention='auto'):
    """
    Parses the Purchase Register Excel/CSV file using the provided column mappings.
    Standardizes and normalizes the data for reconciliation.
    Uses memory-optimized streaming reader for Excel.
    """
    # 1. Load file
    if hasattr(file_path_or_buffer, 'columns') or type(file_path_or_buffer).__name__ == 'DataFrame':
        df_raw = file_path_or_buffer
    elif isinstance(file_path_or_buffer, str) and file_path_or_buffer.lower().endswith('.csv'):
        # For CSV, read chunked or with low_memory flag
        df_raw = pd.read_csv(file_path_or_buffer, low_memory=True)
    else:
        # Check if legacy .xls file
        is_xls = False
        if isinstance(file_path_or_buffer, str) and file_path_or_buffer.lower().endswith('.xls'):
            is_xls = True
        elif hasattr(file_path_or_buffer, 'name') and str(file_path_or_buffer.name).lower().endswith('.xls'):
            is_xls = True
            
        # Check file size (only use read_only streaming for files > 10MB to avoid openpyxl BytesIO zip bugs)
        is_large = False
        if hasattr(file_path_or_buffer, 'size') and file_path_or_buffer.size > 10 * 1024 * 1024:
            is_large = True
            
        if is_xls:
            df_raw = pd.read_excel(file_path_or_buffer, sheet_name=sheet_name)
        elif is_large:
            df_raw = parse_large_excel_streaming(file_path_or_buffer, sheet_name=sheet_name)
        else:
            df_raw = pd.read_excel(file_path_or_buffer, sheet_name=sheet_name)

    # 2. Extract and map fields
    df = pd.DataFrame()
    
    # Map Supplier GSTIN
    gst_col = col_mapping.get('supplier_gstin')
    if gst_col and gst_col in df_raw.columns:
        df['supplier_gstin'] = df_raw[gst_col].astype(str).str.strip().str.upper()
    else:
        df['supplier_gstin'] = 'UNKNOWN'
        
    # Map Document Number
    num_col = col_mapping.get('doc_num')
    if num_col and num_col in df_raw.columns:
        df['doc_num'] = df_raw[num_col].astype(str).str.strip()
        df['clean_doc_num'] = df['doc_num'].apply(clean_invoice_number)
    else:
        raise ValueError("Document/Invoice Number column mapping is required.")
        
    # Map Document Date
    date_col = col_mapping.get('doc_date')
    if date_col and date_col in df_raw.columns:
        df['doc_date'] = df_raw[date_col].apply(parse_date)
    else:
        df['doc_date'] = pd.NaT
        
    # Map Supplier Name
    name_col = col_mapping.get('supplier_name')
    if name_col and name_col in df_raw.columns:
        df['supplier_name'] = df_raw[name_col].fillna("").astype(str).str.strip()
    else:
        df['supplier_name'] = ""
        
    # Map Document Type
    type_col = col_mapping.get('doc_type')
    if type_col and type_col in df_raw.columns:
        df['doc_type_raw'] = df_raw[type_col].fillna("").astype(str).str.strip().str.upper()
    else:
        df['doc_type_raw'] = 'INV'
        
    # Map Place of Supply
    pos_col = col_mapping.get('pos')
    if pos_col and pos_col in df_raw.columns:
        df['pos'] = df_raw[pos_col].fillna("").astype(str).str.strip()
    else:
        df['pos'] = ""

    # Map Reverse Charge
    rc_col = col_mapping.get('rchrg')
    if rc_col and rc_col in df_raw.columns:
        def parse_rc(val):
            val_str = str(val).strip().upper()
            return 'Yes' if val_str in ('Y', 'YES', 'TRUE', '1') else 'No'
        df['rchrg'] = df_raw[rc_col].apply(parse_rc)
    else:
        df['rchrg'] = 'No'

    # Map Monetary values
    val_cols = {
        'taxable_val': col_mapping.get('taxable_val'),
        'igst': col_mapping.get('igst'),
        'cgst': col_mapping.get('cgst'),
        'sgst': col_mapping.get('sgst'),
        'cess': col_mapping.get('cess'),
        'total_val': col_mapping.get('total_val')
    }
    
    for key, col_name in val_cols.items():
        if col_name and col_name in df_raw.columns:
            df[key] = pd.to_numeric(df_raw[col_name].astype(str).replace(r'[\$,₹,]', '', regex=True), errors='coerce').fillna(0.0)
        else:
            df[key] = 0.0

    # Determine standard document type (INV, CRN, DBN)
    def determine_doc_type(row):
        t_raw = str(row.get('doc_type_raw', '')).upper()
        if 'CREDIT' in t_raw or 'CRN' in t_raw or 'CN' == t_raw or 'CDN' in t_raw:
            return 'CRN'
        elif 'DEBIT' in t_raw or 'DBN' in t_raw or 'DN' == t_raw:
            return 'DBN'
        elif row.get('taxable_val') < 0:
            return 'CRN'
        return 'INV'

    df['doc_type'] = df.apply(determine_doc_type, axis=1)

    # Normalize Credit Note values to negative
    def normalize_signs(row):
        is_cn = (row['doc_type'] == 'CRN')
        is_negative = (row['taxable_val'] < 0)
        if is_cn and not is_negative:
            row['taxable_val'] = -abs(row['taxable_val'])
            row['igst'] = -abs(row['igst'])
            row['cgst'] = -abs(row['cgst'])
            row['sgst'] = -abs(row['sgst'])
            row['cess'] = -abs(row['cess'])
            row['total_val'] = -abs(row['total_val'])
        return row

    if credit_note_convention in ('auto', 'negative'):
        df = df.apply(normalize_signs, axis=1)

    # Additional standard fields for Books
    df['is_amended'] = False
    df['original_doc_num'] = None
    df['original_doc_date'] = None
    df['rtn_period'] = None
    df['itc_eligibility'] = 'Eligible'
    df['filing_date'] = pd.NaT
    df['gstr3b_status'] = 'Yes'
    df['section'] = 'Purchase Register'
    df['source'] = 'Books'
    df['source_file'] = 'Purchase Register'

    # Round all monetary values
    for val_col in ['taxable_val', 'igst', 'cgst', 'sgst', 'cess', 'total_val']:
        df[val_col] = df[val_col].round(2)

    # Clean up raw DataFrame references and collect garbage
    del df_raw
    gc.collect()

    return df

def auto_detect_columns(columns):
    """
    Auto-detect columns from a list of column headers using common keywords.
    """
    col_lower = [str(c).lower().strip() for c in columns]
    detected = {}
    
    keywords = {
        'supplier_gstin': ['gstin', 'gst', 'gst no', 'gstin/uin', 'supplier gst', 'ctin', 'party gst', 'vendor gstin'],
        'supplier_name': ['supplier name', 'vendor name', 'party name', 'name', 'supplier_name', 'party_name', 'lglname', 'legal name', 'vendor name'],
        'doc_num': ['invoice number', 'invoice no', 'inv no', 'bill no', 'voucher no', 'document number', 'doc no', 'invoice_no', 'inv_num', 'inum', 'document no', 'bill number'],
        'doc_date': ['invoice date', 'date', 'inv date', 'bill date', 'voucher date', 'doc date', 'invoice_dt', 'idt', 'invoice_date', 'bill date'],
        'taxable_val': ['taxable value', 'taxable amount', 'taxable amt', 'taxable val', 'assessable value', 'taxable_value', 'taxable_amt', 'txval', 'purchase value'],
        'igst': ['igst', 'integrated tax', 'igst amount', 'igst amt', 'igst_amt', 'igst_val'],
        'cgst': ['cgst', 'central tax', 'cgst amount', 'cgst amt', 'cgst_amt', 'cgst_val'],
        'sgst': ['sgst', 'state tax', 'sgst amount', 'sgst amt', 'sgst_amt', 'sgst_val', 'utgst', 'utgst amt', 'utgst_amount'],
        'cess': ['cess', 'cess amount', 'cess amt', 'cess_amt', 'cess_val', 'cec'],
        'total_val': ['total value', 'invoice value', 'total amount', 'inv value', 'bill amount', 'invoice_val', 'val', 'total_amt', 'gross value', 'invoice amount'],
        'doc_type': ['document type', 'doc type', 'voucher type', 'vtype', 'type', 'doc_type', 'voucher name'],
        'pos': ['pos', 'place of supply', 'place_of_supply', 'state code', 'supply state'],
        'rchrg': ['rchrg', 'reverse charge', 'rcm', 'rc', 'reverse_charge']
    }
    
    for field, terms in keywords.items():
        matched = False
        for term in terms:
            for idx, col in enumerate(col_lower):
                if term == col or (len(term) >= 3 and term in col):
                    detected[field] = columns[idx]
                    matched = True
                    break
            if matched:
                break
                
    return detected
